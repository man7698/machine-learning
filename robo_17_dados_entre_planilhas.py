# -*- coding: utf-8 -*-
"""
Created on Tue Oct 31 16:52:47 2023

@author: Matheus Almeida
"""


from openpyxl import load_workbook
from openpyxl import Workbook
import os

caminho_nome_arquivo = "C:\\Users\Matheus Almeida\Desktop\Robos Python\Planilhas\\DadosSistema.xlsx"
arquivo = load_workbook(filename=caminho_nome_arquivo)

planilha_selecionada = arquivo['Dados']

criandonovoarquivo = Workbook()

novaplanilha = criandonovoarquivo.active

for linha in range(1, len(planilha_selecionada['A']) + 1):
    for coluna in range(1,10):
        novaplanilha.cell(row=linha, column= coluna).value = planilha_selecionada.cell(row=linha, column= coluna).value 

novaplanilha.totle = "Dados Funcionario"

criandonovoarquivo.create_sheet('Resumo')

caminho_nova_planilha = "C:\\Users\Matheus Almeida\Desktop\Robos Python\Planilhas\\RelatorioSitema.xlsx"


arquivo.save(filename=caminho_nova_planilha)

os.startfile(caminho_nova_planilha)