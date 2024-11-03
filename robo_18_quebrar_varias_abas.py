from openpyxl import load_workbook
import os

nome_arquivo = "C:\\Users\Matheus Almeida\Desktop\Robos Python\Planilhas\\Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

#Seleciona a sheet de Dados
sheet_selecionada = planilha_aberta['Dados']

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    nomeAtual = sheet_selecionada['A%s' % linha].value
    
    if nomeNovo == nomeAtual:
        
        linhaSheetQuebra = len(sheet_selecionada2['A']) + 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)
                    
        #Preenche os dados
        sheet_selecionada2[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2[celulaColunaC] = sheet_selecionada['C%s' % linha].value
        
    else:
        
        #Cria uma nova sheet com o nome do funcionario
        sheet_resumo = planilha_aberta.create_sheet(title=nomeAtual)
        
        #Seleciona a sheet que foi criada
        sheet_selecionada2 = planilha_aberta[nomeAtual]
        
        #Adiciona o nome do funcionario que esta na linha que o código está passando
        nomeNovo = sheet_selecionada['A%s' % linha].value
        
        #Coloca os titulos
        sheet_selecionada2['A1'] = "Vendedor"
        sheet_selecionada2['B1'] = "Produtos"
        sheet_selecionada2['C1'] = "Vendas"
        
        #Preenche as informações na segunda linha
        sheet_selecionada2['A2'] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2['B2'] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2['C2'] = sheet_selecionada['C%s' % linha].value
        
        
#Salva as alterações feitas na planilha
planilha_aberta.save(filename=nome_arquivo)

#Abre o arquivo com as mudanças
os.startfile(nome_arquivo)
