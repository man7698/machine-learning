from openpyxl import load_workbook
from openpyxl import Workbook
import os

nome_arquivo = "C:\\Users\Matheus Almeida\Desktop\Robos Python\Planilhas\\Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

#Seleciona a sheet de Dados
sheet_selecionada = planilha_aberta['Dados']

criandoNovoArquivoExcel = Workbook()

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    nomeAtual = sheet_selecionada['A%s' % linha].value
    
    if nomeNovo == nomeAtual:
        
        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)
                    
        #Preenche os dados
        selecionaSheetVendasNovaPlanilha[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaC] = sheet_selecionada['C%s' % linha].value
        
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)
        
    else:
        
        #Adiciona o nome do funcionario que esta na linha que o código está passando
        nomeNovo = sheet_selecionada['A%s' % linha].value
        
        nova_planilha = criandoNovoArquivoExcel.active
        
        nova_planilha.title = "Vendas"
        
        caminhoNovaPlanilha = "C:\\Users\Matheus Almeida\Desktop\Robos Python\Planilhas\\" + sheet_selecionada['A%s' % linha].value + ".xlsx"
        
        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']
        
        #Coloca os titulos
        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"
        
        #Preenche as informações na segunda linha
        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value
        
        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)
        
        
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)
        
        